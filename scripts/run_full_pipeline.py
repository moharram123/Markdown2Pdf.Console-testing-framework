from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent.parent

CONTROL_DIR = BASE_DIR / "data" / "llm-generated" / "control"
REGRESSION_DIR = BASE_DIR / "data" / "llm-generated" / "regressions"
PDF_DIR = BASE_DIR / "results" / "generated-pdfs" / "llm"
METRICS_DIR = BASE_DIR / "results" / "metrics"

OUTPUT_FILE = METRICS_DIR / "llm_metrics.xlsx"

METRICS_DIR.mkdir(parents=True, exist_ok=True)


def count_files(folder: Path, pattern: str) -> int:
    if not folder.exists():
        return 0
    return len(list(folder.glob(pattern)))


def style_header(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = min(max_length + 3, 35)


def add_borders(ws):
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def create_summary_sheet(wb, metrics):
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Metric", "Value"])
    ws.append(["Control Markdown Files", metrics["control_files"]])
    ws.append(["Regression Markdown Files", metrics["regression_files"]])
    ws.append(["Generated PDFs", metrics["pdf_files"]])
    ws.append(["Total Pytest Cases", metrics["pytest_total"]])
    ws.append(["Passed Pytest Cases", metrics["pytest_passed"]])
    ws.append(["Failed Pytest Cases", metrics["pytest_failed"]])
    ws.append(["Regression Cases Detected", metrics["regression_detected"]])
    ws.append(["Regression Cases Missed", metrics["regression_missed"]])
    ws.append(["Regression Detection Rate (%)", metrics["detection_rate"]])
    ws.append(["False Positive Rate (%)", metrics["false_positive_rate"]])
    ws.append(["Accuracy (%)", metrics["accuracy"]])

    style_header(ws)
    add_borders(ws)
    auto_fit_columns(ws)

    chart = BarChart()
    chart.title = "LLM Pipeline Metrics Overview"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Metric"

    data = Reference(ws, min_col=2, min_row=1, max_row=12)
    categories = Reference(ws, min_col=1, min_row=2, max_row=12)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 18

    ws.add_chart(chart, "D2")


def create_accuracy_sheet(wb, metrics):
    ws = wb.create_sheet("Accuracy Analysis")

    ws.append(["Category", "Value"])
    ws.append(["Passed", metrics["pytest_passed"]])
    ws.append(["Failed", metrics["pytest_failed"]])

    style_header(ws)
    add_borders(ws)
    auto_fit_columns(ws)

    chart = PieChart()
    chart.title = "Overall Test Accuracy"

    data = Reference(ws, min_col=2, min_row=2, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)

    chart.add_data(data)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 12

    ws.add_chart(chart, "D2")


def create_detection_sheet(wb, metrics):
    ws = wb.create_sheet("Detection Rate")

    ws.append(["Regression Result", "Count"])
    ws.append(["Detected Regressions", metrics["regression_detected"]])
    ws.append(["Missed Regressions", metrics["regression_missed"]])

    style_header(ws)
    add_borders(ws)
    auto_fit_columns(ws)

    chart = BarChart()
    chart.title = "Regression Detection Rate"
    chart.y_axis.title = "Number of Cases"
    chart.x_axis.title = "Regression Result"

    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    categories = Reference(ws, min_col=1, min_row=2, max_row=3)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 14

    ws.add_chart(chart, "D2")


def create_false_positive_sheet(wb, metrics):
    ws = wb.create_sheet("False Positive Rate")

    ws.append(["Metric", "Value"])
    ws.append(["Control Cases", metrics["control_files"]])
    ws.append(["False Positives", metrics["false_positives"]])
    ws.append(["False Positive Rate (%)", metrics["false_positive_rate"]])

    style_header(ws)
    add_borders(ws)
    auto_fit_columns(ws)

    chart = BarChart()
    chart.title = "False Positive Analysis"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Metric"

    data = Reference(ws, min_col=2, min_row=1, max_row=4)
    categories = Reference(ws, min_col=1, min_row=2, max_row=4)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 9
    chart.width = 14

    ws.add_chart(chart, "D2")


def main():
    control_files = count_files(CONTROL_DIR, "*.md")
    regression_files = count_files(REGRESSION_DIR, "*.md")
    pdf_files = count_files(PDF_DIR, "*.pdf")

    # Update these values if your pytest result changes.
    pytest_total = 34
    pytest_passed = 34
    pytest_failed = 0

    # Update these values based on the latest regression detection result.
    regression_detected = 33
    regression_missed = max(regression_files - regression_detected, 0)

    false_positives = 0

    detection_rate = round((regression_detected / regression_files) * 100, 2) if regression_files else 0
    false_positive_rate = round((false_positives / control_files) * 100, 2) if control_files else 0
    accuracy = round((pytest_passed / pytest_total) * 100, 2) if pytest_total else 0

    metrics = {
        "control_files": control_files,
        "regression_files": regression_files,
        "pdf_files": pdf_files,
        "pytest_total": pytest_total,
        "pytest_passed": pytest_passed,
        "pytest_failed": pytest_failed,
        "regression_detected": regression_detected,
        "regression_missed": regression_missed,
        "detection_rate": detection_rate,
        "false_positives": false_positives,
        "false_positive_rate": false_positive_rate,
        "accuracy": accuracy,
    }

    wb = Workbook()

    create_summary_sheet(wb, metrics)
    create_accuracy_sheet(wb, metrics)
    create_detection_sheet(wb, metrics)
    create_false_positive_sheet(wb, metrics)

    wb.save(OUTPUT_FILE)

    print(f"Excel metrics report saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()